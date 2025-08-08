import openpyxl
import os
import time
from tqdm import tqdm


def process_excel_readonly(input_path):
    # 原文件只读打开
    wb_in = openpyxl.load_workbook(input_path, read_only=True)

    # 创建新工作簿用于写入
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # 删除默认工作表

    for sheet_in in wb_in.worksheets:
        # 新建对应名称的 sheet
        sheet_out = wb_out.create_sheet(title=sheet_in.title)
        rows = list(sheet_in.iter_rows())

        for row in tqdm(rows, desc=f"处理工作表: {sheet_in.title}"):
            new_row = []
            for cell in row:
                value = str(cell.value).strip() if cell.value is not None else ""
                if "#" in value:
                    before_hash = value.split("#", 1)[0].strip()
                    if len(before_hash) > 10:
                        value = ""  # 清空内容
                new_row.append(value)
            sheet_out.append(new_row)  # 写入新 sheet

    # 输出路径
    base_dir = os.path.dirname(input_path)
    file_name = os.path.splitext(os.path.basename(input_path))[0]
    timestamp = time.strftime("%Y%m%d%H%M%S")
    output_path = os.path.join(base_dir, f"{file_name}_{timestamp}.xlsx")

    wb_out.save(output_path)
    print(f"\n✅ 处理完成，输出文件：{output_path}")


if __name__ == "__main__":
    # 方法一：手动写路径（推荐）
    input_excel = r"D:\桌面\8spa\文本库_2025-08-08 15_14_16.xlsx"

    # 方法二：命令行输入路径（可选）
    # input_excel = input("请输入 Excel 文件路径：").strip().strip('"')

    if not os.path.exists(input_excel):
        print("❌ 输入的文件路径不存在！")
    else:
        process_excel_readonly(input_excel)
